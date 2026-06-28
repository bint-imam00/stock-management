"""Rapports : état du stock, mouvements filtrés, top ventes, export Excel."""
from datetime import datetime
from io import BytesIO

from flask import Blueprint, render_template, request, send_file
from flask_login import login_required
from sqlalchemy import func
from openpyxl import Workbook

from extensions import db
from models import Product, Category, StockIn, StockOut

bp = Blueprint("reports", __name__, url_prefix="/reports")


def _parse_period():
    """Lit les paramètres ?from=YYYY-MM-DD&to=YYYY-MM-DD."""
    date_from = request.args.get("from", "").strip()
    date_to = request.args.get("to", "").strip()
    dt_from = dt_to = None
    if date_from:
        try:
            dt_from = datetime.strptime(date_from, "%Y-%m-%d")
        except ValueError:
            dt_from = None
    if date_to:
        try:
            dt_to = datetime.strptime(date_to, "%Y-%m-%d")
            dt_to = dt_to.replace(hour=23, minute=59, second=59)
        except ValueError:
            dt_to = None
    return dt_from, dt_to, date_from, date_to


@bp.route("/")
@login_required
def index():
    return render_template("reports/index.html")


@bp.route("/stock")
@login_required
def stock():
    """État global du stock."""
    items = Product.query.order_by(Product.name).all()
    total_value = sum(p.stock_value for p in items)
    return render_template("reports/stock.html", items=items, total_value=total_value)


@bp.route("/movements")
@login_required
def movements():
    """Historique des mouvements (entrées + sorties) filtrable par période."""
    dt_from, dt_to, date_from, date_to = _parse_period()

    q_in = StockIn.query
    q_out = StockOut.query
    if dt_from:
        q_in = q_in.filter(StockIn.created_at >= dt_from)
        q_out = q_out.filter(StockOut.created_at >= dt_from)
    if dt_to:
        q_in = q_in.filter(StockIn.created_at <= dt_to)
        q_out = q_out.filter(StockOut.created_at <= dt_to)

    stock_ins = q_in.order_by(StockIn.created_at.desc()).all()
    stock_outs = q_out.order_by(StockOut.created_at.desc()).all()

    return render_template(
        "reports/movements.html",
        stock_ins=stock_ins,
        stock_outs=stock_outs,
        date_from=date_from,
        date_to=date_to,
    )


@bp.route("/top-sold")
@login_required
def top_sold():
    """Produits les plus vendus, filtrable par période."""
    dt_from, dt_to, date_from, date_to = _parse_period()

    query = (
        db.session.query(
            Product.name,
            Category.name.label("category"),
            func.coalesce(func.sum(StockOut.quantity), 0).label("total"),
        )
        .join(StockOut, StockOut.product_id == Product.id)
        .join(Category, Category.id == Product.category_id)
    )
    if dt_from:
        query = query.filter(StockOut.created_at >= dt_from)
    if dt_to:
        query = query.filter(StockOut.created_at <= dt_to)
    rows = query.group_by(Product.id).order_by(func.sum(StockOut.quantity).desc()).all()

    return render_template(
        "reports/top_sold.html",
        rows=rows,
        date_from=date_from,
        date_to=date_to,
    )


@bp.route("/export/stock.xlsx")
@login_required
def export_stock():
    """Export Excel de l'état du stock."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Stock"
    ws.append(["Produit", "SKU", "Catégorie", "Prix unitaire", "Quantité", "Stock min", "Valeur"])
    for p in Product.query.order_by(Product.name).all():
        ws.append([
            p.name,
            p.sku or "",
            p.category.name if p.category else "",
            p.price,
            p.quantity,
            p.min_stock,
            p.stock_value,
        ])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name=f"stock_{datetime.now().strftime('%Y%m%d')}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@bp.route("/export/movements.xlsx")
@login_required
def export_movements():
    """Export Excel des mouvements sur la période demandée."""
    dt_from, dt_to, _, _ = _parse_period()

    q_in = StockIn.query
    q_out = StockOut.query
    if dt_from:
        q_in = q_in.filter(StockIn.created_at >= dt_from)
        q_out = q_out.filter(StockOut.created_at >= dt_from)
    if dt_to:
        q_in = q_in.filter(StockIn.created_at <= dt_to)
        q_out = q_out.filter(StockOut.created_at <= dt_to)

    wb = Workbook()
    ws_in = wb.active
    ws_in.title = "Entrées"
    ws_in.append(["Date", "Produit", "Fournisseur", "Quantité", "Prix unitaire", "Note"])
    for e in q_in.order_by(StockIn.created_at.desc()).all():
        ws_in.append([
            e.created_at.strftime("%Y-%m-%d %H:%M"),
            e.product.name if e.product else "",
            e.supplier.name if e.supplier else "",
            e.quantity,
            e.unit_price or 0,
            e.note or "",
        ])

    ws_out = wb.create_sheet("Sorties")
    ws_out.append(["Date", "Produit", "Quantité", "Motif"])
    for s in q_out.order_by(StockOut.created_at.desc()).all():
        ws_out.append([
            s.created_at.strftime("%Y-%m-%d %H:%M"),
            s.product.name if s.product else "",
            s.quantity,
            s.reason or "",
        ])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name=f"mouvements_{datetime.now().strftime('%Y%m%d')}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
